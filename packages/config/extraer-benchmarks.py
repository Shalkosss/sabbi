"""
Extrae los benchmarks a precision completa desde la macro de referencia.

Entrada : reference/Portfolio Modificado (version 1).xlsm, hoja `Data`
Salida  : packages/config/benchmarks.v3.json

Por que existe este script
--------------------------
`sabbi_configuracion_v2.json` trae los mismos pesos redondeados a cuatro
decimales. Sobre el caso Ana Tumi ese redondeo desplaza la base de
redistribucion 6,502.88 USD y ningun golden test puede pasar. La hoja `Data` es
la fuente autoritativa y este script la vuelca sin perder digitos.

La hoja codifica la jerarquia con sangria: una fila sin espacios al inicio es
una clase; una con sangria es un producto de la clase anterior. El peso de cada
producto esta expresado sobre el patrimonio total, no sobre su clase.
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parents[2]
ORIGEN = RAIZ / "reference" / "Portfolio Modificado (version 1).xlsm"
DESTINO = Path(__file__).resolve().parent / "benchmarks.v3.json"

VERSION = 3
COL_PRIMER_PERFIL = 3
FILA_ENCABEZADO = 2

# Nombre en la hoja -> clave del motor. La hoja escribe "Imobiliario" sin la n.
CLASES = {
    "Imobiliario Directo": "inm",
    "Inmobiliario Directo": "inm",
    "Mercados Publicos - Fijo": "fijo",
    "Mercados Publicos - Variable": "variable",
    "Mercados Privados": "privados",
    "Cash": "cash",
}

# Tolerancia al validar que cada perfil sume 1.
TOL = 1e-9


def leer() -> dict:
    libro = openpyxl.load_workbook(ORIGEN, data_only=True)
    hoja = libro["Data"]

    perfiles: list[str] = []
    col = COL_PRIMER_PERFIL
    while (valor := hoja.cell(FILA_ENCABEZADO, col).value) is not None:
        perfiles.append(str(valor).strip())
        col += 1
    if not perfiles:
        raise SystemExit("No encuentro los perfiles en la fila de encabezado de Data.")

    clases: dict[str, dict] = {}
    clase_actual: str | None = None

    fila = FILA_ENCABEZADO + 1
    while True:
        crudo = hoja.cell(fila, 2).value
        if crudo is None:
            fila += 1
            if fila > hoja.max_row:
                break
            continue

        etiqueta = str(crudo)
        nombre = etiqueta.strip()
        if nombre.upper() == "TOTAL":
            break

        pesos = {
            perfil: _numero(hoja.cell(fila, COL_PRIMER_PERFIL + i).value)
            for i, perfil in enumerate(perfiles)
        }

        # sangria = producto de la clase anterior
        if etiqueta.startswith(" ") or etiqueta.startswith("\t"):
            if clase_actual is None:
                raise SystemExit(f"Fila {fila}: producto {nombre!r} sin clase padre.")
            clases[clase_actual]["productos"].append({"nombre": nombre, "pesos": pesos})
        else:
            clave = CLASES.get(nombre)
            if clave is None:
                raise SystemExit(
                    f"Fila {fila}: la clase {nombre!r} no esta en el mapa. "
                    "Agregala a CLASES antes de continuar."
                )
            clase_actual = clave
            clases[clave] = {"nombre": nombre, "pesos": pesos, "productos": []}

        fila += 1

    _validar(clases, perfiles)
    return {
        "version": VERSION,
        "origen": ORIGEN.name,
        "hoja": "Data",
        "nota": (
            "Pesos a precision completa. No usar los de sabbi_configuracion_v2.json: "
            "estan redondeados a cuatro decimales y mueven la base de redistribucion."
        ),
        "perfiles": perfiles,
        "clases": clases,
    }


def _numero(valor) -> float:
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    raise SystemExit(f"Peso no numerico en la hoja: {valor!r}")


def _validar(clases: dict, perfiles: list[str]) -> None:
    for perfil in perfiles:
        total = sum(c["pesos"][perfil] for c in clases.values())
        if abs(total - 1.0) > 1e-6:
            raise SystemExit(f"Los pesos de {perfil!r} suman {total!r}, deberian sumar 1.")

    # Un producto nunca puede pesar mas que su clase.
    for clave, clase in clases.items():
        if not clase["productos"]:
            continue
        for perfil in perfiles:
            hijos = sum(p["pesos"][perfil] for p in clase["productos"])
            padre = clase["pesos"][perfil]
            if hijos > padre + 1e-9:
                raise SystemExit(
                    f"{clave}/{perfil}: los productos suman {hijos} y la clase vale {padre}."
                )


def main() -> None:
    datos = leer()
    DESTINO.write_text(json.dumps(datos, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"escrito: {DESTINO.relative_to(RAIZ)}")
    print(f"perfiles: {len(datos['perfiles'])}  clases: {len(datos['clases'])}")
    for clave, clase in datos["clases"].items():
        print(f"  {clave:<9} {len(clase['productos'])} productos")
    moderado = {k: v["pesos"]["Moderado"] for k, v in datos["clases"].items()}
    print(f"  Moderado suma {sum(moderado.values())!r}")


if __name__ == "__main__":
    main()
